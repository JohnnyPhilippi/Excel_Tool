import openpyxl
from googletrans import Translator
import streamlit as st

def translate_excel(input_file, source_lang='en', target_lang='de'):
    """
    Translates the contents of an Excel file from one language to another while preserving formatting.

    Args:
        input_file: Uploaded Excel file.
        source_lang (str): Source language code (e.g., 'en').
        target_lang (str): Target language code (e.g., 'de').

    Returns:
        Translated Excel workbook.
    """
    workbook = openpyxl.load_workbook(input_file)
    translator = Translator()

    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]

        for row in sheet.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    try:
                        translated_text = translator.translate(cell.value, src=source_lang, dest=target_lang).text
                        cell.value = translated_text
                    except Exception as e:
                        st.error(f"Error translating cell {cell.coordinate}: {e}")
    return workbook

# Streamlit App
st.title("Excel Übersetzungstool")
st.write("Lade eine Excel-Datei hoch und übersetze sie zwischen Englisch und Deutsch.")

# Datei-Upload
uploaded_file = st.file_uploader("Lade eine Excel-Datei hoch", type=["xlsx"])
source_language = st.selectbox("Ausgangssprache", ["Englisch", "Deutsch"])
target_language = st.selectbox("Zielsprache", ["Deutsch", "Englisch"])

if st.button("Übersetzen"):
    if uploaded_file:
        source_lang = 'en' if source_language == "Englisch" else 'de'
        target_lang = 'de' if target_language == "Deutsch" else 'en'

        # Übersetzen
        translated_workbook = translate_excel(uploaded_file, source_lang, target_lang)

        # Datei speichern und zum Download bereitstellen
        from io import BytesIO
        translated_file = BytesIO()
        translated_workbook.save(translated_file)
        translated_file.seek(0)

        st.download_button(
            label="Übersetzte Datei herunterladen",
            data=translated_file,
            file_name="translated_file.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    else:
        st.error("Bitte lade zuerst eine Datei hoch.")
